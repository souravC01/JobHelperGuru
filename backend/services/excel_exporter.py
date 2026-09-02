import io
from typing import List
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from backend.models import Application, ApplicationStatus


class ExcelExporter:
    STATUS_COLORS = {
        ApplicationStatus.WISHLIST: "F1F5F9",     # Slate Light
        ApplicationStatus.APPLIED: "DBEAFE",      # Soft Blue
        ApplicationStatus.INTERVIEWING: "FEF3C7", # Soft Amber
        ApplicationStatus.OFFERED: "DCFCE7",      # Soft Emerald
        ApplicationStatus.REJECTED: "FFE4E6",     # Soft Rose
        ApplicationStatus.ARCHIVED: "E2E8F0",     # Muted Gray
    }

    def export_workbook(self, applications: List[Application]) -> bytes:
        wb = openpyxl.Workbook()

        # Styles
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        data_font = Font(name="Segoe UI", size=10)
        bold_font = Font(name="Segoe UI", size=10, bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        url_font = Font(name="Segoe UI", size=10, color="2563EB", underline="single")

        thin_border_side = Side(border_style="thin", color="E2E8F0")
        cell_border = Border(
            left=thin_border_side,
            right=thin_border_side,
            top=thin_border_side,
            bottom=thin_border_side,
        )

        # -------------------------------------------------------------
        # Sheet 1: Applications Tracker
        # -------------------------------------------------------------
        ws1 = wb.active
        ws1.title = "Applications Tracker"
        ws1.freeze_panes = "A2"
        ws1.row_dimensions[1].height = 28

        headers1 = [
            "Date Added",
            "Company",
            "Role / Title",
            "Status",
            "Location",
            "Salary Range",
            "Job Link",
            "Application Date",
            "Follow-Up Date",
            "Top ATS Keywords",
            "Required Skills",
            "Notes",
        ]

        for col_idx, header in enumerate(headers1, 1):
            cell = ws1.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        for row_idx, app in enumerate(applications, 2):
            ws1.row_dimensions[row_idx].height = 22

            status_val = app.status.value if isinstance(app.status, ApplicationStatus) else str(app.status)
            status_enum = ApplicationStatus(status_val) if status_val in [s.value for s in ApplicationStatus] else ApplicationStatus.WISHLIST

            row_data = [
                app.date_added,
                app.company,
                app.role,
                status_val,
                app.location or "Unknown",
                app.salary or "Not specified",
                app.url or "",
                app.application_date or "",
                app.follow_up_date or "",
                ", ".join(app.ats_keywords[:8]),
                ", ".join(app.required_skills[:8]),
                app.notes or "",
            ]

            for col_idx, val in enumerate(row_data, 1):
                cell = ws1.cell(row=row_idx, column=col_idx, value=val)
                cell.font = data_font
                cell.border = cell_border
                cell.alignment = left_align

                # Company & Role Bold
                if col_idx in [2, 3]:
                    cell.font = bold_font

                # Center Dates & Status
                if col_idx in [1, 4, 8, 9]:
                    cell.alignment = center_align

                # Status color fill
                if col_idx == 4:
                    hex_color = self.STATUS_COLORS.get(status_enum, "F1F5F9")
                    cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

                # Job link hyperlink
                if col_idx == 7 and val and (val.startswith("http://") or val.startswith("https://")):
                    cell.font = url_font
                    cell.hyperlink = val

        # Auto-adjust column widths for Sheet 1
        self._auto_fit_columns(ws1, min_width=12, max_width=45)

        # -------------------------------------------------------------
        # Sheet 2: Skills & ATS Keywords
        # -------------------------------------------------------------
        ws2 = wb.create_sheet(title="Skills & ATS Keywords")
        ws2.freeze_panes = "A2"
        ws2.row_dimensions[1].height = 28

        headers2 = [
            "Company",
            "Role",
            "Required Skills (Must Haves)",
            "Top ATS Keywords",
            "Notes / Strategy",
        ]

        for col_idx, header in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        for row_idx, app in enumerate(applications, 2):
            ws2.row_dimensions[row_idx].height = 24
            row_data = [
                app.company,
                app.role,
                ", ".join(app.required_skills),
                ", ".join(app.ats_keywords),
                app.notes or "",
            ]
            for col_idx, val in enumerate(row_data, 1):
                cell = ws2.cell(row=row_idx, column=col_idx, value=val)
                cell.font = data_font
                cell.border = cell_border
                cell.alignment = left_align
                if col_idx in [1, 2]:
                    cell.font = bold_font

        self._auto_fit_columns(ws2, min_width=15, max_width=50)

        # Write to memory buffer
        bio = io.BytesIO()
        wb.save(bio)
        return bio.getvalue()

    def _auto_fit_columns(self, ws, min_width: int = 12, max_width: int = 45):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or "")
                if val:
                    # If multiple lines, take max line length
                    line_lens = [len(line) for line in val.split("\n")]
                    max_len = max(max_len, max(line_lens))
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, min_width), max_width)
