import io
import openpyxl
from fastapi.testclient import TestClient
from backend.main import app

client = TestClient(app)


def get_auth_headers():
    res = client.post(
        "/api/auth/register",
        json={
            "email": "e2e_tester@example.com",
            "password": "Password123!",
            "name": "E2E Tester",
        },
    )
    if res.status_code == 200:
        token = res.json()["token"]
    else:
        res_login = client.post(
            "/api/auth/login",
            json={"email": "e2e_tester@example.com", "password": "Password123!"},
        )
        token = res_login.json()["token"]
    return {"Authorization": f"Bearer {token}"}


def test_full_user_journey_e2e():
    headers = get_auth_headers()

    # 1. Analyze Job Description
    stripe_jd = """
    Senior Backend Engineer at Stripe
    Location: San Francisco, CA (Hybrid)
    Salary: $160,000 - $210,000
    Requirements:
    - 5+ years of experience with Python, FastAPI, and PostgreSQL.
    - Strong knowledge of Docker, Kubernetes, and AWS microservices.
    - Excellent communication and Agile collaboration.
    Bonus:
    - Experience with Kafka, Redis, and GraphQL.
    """
    res = client.post("/api/jobs/analyze", json={"text": stripe_jd})
    assert res.status_code == 200
    job_data = res.json()

    assert job_data["company"] == "Stripe" or "Stripe" in job_data["raw_text"]
    assert "Python" in job_data["required_skills"] or "Python" in job_data["tech_stack"]
    assert "PostgreSQL" in job_data["required_skills"] or "PostgreSQL" in job_data["tech_stack"]
    assert "$160,000 - $210,000" in job_data["salary_range"]
    assert job_data["work_mode"] == "Hybrid"
    assert len(job_data["ats_keywords"]) >= 5

    # 2. Save Analyzed Job to Applications Tracker
    create_app_res = client.post(
        "/api/applications",
        json={
            "company": "Stripe",
            "role": "Senior Backend Engineer",
            "status": "Wishlist",
            "location": "San Francisco, CA (Hybrid)",
            "salary": "$160,000 - $210,000",
            "required_skills": job_data["required_skills"],
            "ats_keywords": job_data["ats_keywords"],
            "notes": job_data["summary"],
        },
        headers=headers,
    )
    assert create_app_res.status_code == 200
    app_record = create_app_res.json()
    app_id = app_record["id"]
    assert app_record["status"] == "Wishlist"

    # 3. Add Resumes to Library
    r1 = client.post(
        "/api/resumes",
        json={
            "name": "Python Backend Engineer",
            "content": "Senior Software Engineer with 4 years building Python, FastAPI, Docker, and PostgreSQL microservices. Implemented database schema migrations and automated testing with pytest.",
        },
        headers=headers,
    ).json()

    r2 = client.post(
        "/api/resumes",
        json={
            "name": "Frontend React Engineer",
            "content": "Frontend developer proficient in React, Tailwind, Next.js, HTML5, CSS3, Figma design systems.",
        },
        headers=headers,
    ).json()

    # 4. Rank Resumes Against Job (Best-Fit Matcher)
    rank_res = client.post(
        "/api/resumes/match",
        json={
            "job": job_data["analysis"],
            "resumes": [r1, r2],
        },
        headers=headers,
    )
    assert rank_res.status_code == 200
    ranked = rank_res.json()
    assert len(ranked) == 2
    # The Python Backend resume must be ranked #1
    assert ranked[0]["resume_name"] == "Python Backend Engineer"
    assert ranked[0]["is_best_fit"] is True
    assert ranked[0]["match_score"] > ranked[1]["match_score"]
    assert "Python" in ranked[0]["matched_keywords"]
    assert "Kafka" in ranked[0]["missing_keywords"] or "Kubernetes" in ranked[0]["missing_keywords"]

    # 5. Optimize Bullet Point for Missing Keyword (Kafka) using BulletSkill 2.0
    opt_res = client.post(
        "/api/resumes/optimize-bullet",
        json={
            "target_job_title": "Senior Backend Engineer",
            "section_type": "project",
            "target_keyword": "Kafka",
            "existing_bullet": "Built microservices for order processing.",
            "evidence_context": [r1["content"]],
        },
        headers=headers,
    )
    assert opt_res.status_code == 200
    opt_data = opt_res.json()
    assert opt_data["target_keyword"] == "Kafka"
    assert opt_data["claim_status"] == "unverified_skill"
    assert opt_data["requires_confirmation"] is True
    assert "Kafka" in opt_data["warning"]
    assert len(opt_data["alternatives"]) == 3
    # Check Candidate A, B, C variants
    variants = [a["variant_name"] for a in opt_data["alternatives"]]
    assert any("Candidate A" in v for v in variants)
    assert any("Candidate B" in v for v in variants)
    assert any("Candidate C" in v for v in variants)
    # Check What + How + Result
    for alt in opt_data["alternatives"]:
        assert alt["what"] == "Kafka"
        assert alt["how"] != ""
        assert alt["result_or_reason"] != ""

    # 6. Generate Tailored Cover Letter Pitch
    outreach_res = client.post(
        "/api/resumes/generate-outreach",
        json={
            "job": job_data["analysis"],
            "resume_id": r1["id"],
        },
        headers=headers,
    )
    assert outreach_res.status_code == 200
    outreach = outreach_res.json()
    assert len(outreach["subject_line"]) > 5
    assert len(outreach["cover_letter_pitch"]) > 80
    assert len(outreach["connection_note"]) > 20

    # 7. Update Application Status to 'Applied'
    update_res = client.patch(
        f"/api/applications/{app_id}",
        json={
            "status": "Applied",
            "follow_up_date": "2026-09-09",
            "notes": "Applied on company site. Optimized resume with Kafka bullet.",
        },
        headers=headers,
    )
    assert update_res.status_code == 200
    assert update_res.json()["status"] == "Applied"
    assert update_res.json()["follow_up_date"] == "2026-09-09"

    # 8. Download and Validate Excel Workbook (.xlsx)
    excel_res = client.get("/api/export/excel", headers=headers)
    assert excel_res.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(excel_res.content))
    assert "Applications Tracker" in wb.sheetnames
    assert "Skills & ATS Keywords" in wb.sheetnames

    ws1 = wb["Applications Tracker"]
    # Verify Stripe is in the excel rows
    found_stripe = any(ws1.cell(row=r, column=2).value == "Stripe" for r in range(2, ws1.max_row + 1))
    assert found_stripe

    # Clean up
    client.delete(f"/api/applications/{app_id}", headers=headers)
    client.delete(f"/api/resumes/{r1['id']}", headers=headers)
    client.delete(f"/api/resumes/{r2['id']}", headers=headers)
