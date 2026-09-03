import io
import openpyxl
from backend.services.excel_exporter import ExcelExporter
from backend.models import Application, ApplicationStatus

def test_generate_excel_workbook():
    exporter = ExcelExporter()
    apps = [
        Application(
            id="app-1",
            company="Google",
            role="Staff Software Engineer",
            status=ApplicationStatus.INTERVIEWING,
            location="Mountain View, CA",
            salary="$220,000 - $280,000",
            url="https://careers.google.com/jobs/1",
            required_skills=["Python", "Distributed Systems", "Kubernetes"],
            ats_keywords=["GCP", "High Throughput", "Architecture"],
            date_added="2026-09-01",
            application_date="2026-09-02",
            follow_up_date="2026-09-09",
            notes="Completed screening call with recruiter"
        ),
        Application(
            id="app-2",
            company="Netflix",
            role="Senior Backend Engineer",
            status=ApplicationStatus.APPLIED,
            location="Los Gatos, CA",
            salary="$200,000 - $250,000",
            url="https://netflix.com/jobs/2",
            required_skills=["Java", "Spring Boot", "Kafka"],
            ats_keywords=["Microservices", "Event-Driven"],
            date_added="2026-09-02",
            application_date="2026-09-02",
            follow_up_date="2026-09-10",
            notes="Applied via referral"
        )
    ]
    excel_bytes = exporter.export_workbook(apps)
    assert len(excel_bytes) > 2000
    
    # Validate with openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    assert "Applications Tracker" in wb.sheetnames
    assert "Skills & ATS Keywords" in wb.sheetnames
    
    ws1 = wb["Applications Tracker"]
    assert ws1.cell(row=1, column=2).value == "Company"
    assert ws1.cell(row=2, column=2).value == "Google"
    assert ws1.cell(row=2, column=4).value == "Interviewing"
    assert ws1.cell(row=3, column=2).value == "Netflix"
    assert ws1.cell(row=3, column=4).value == "Applied"
    
    # Check hyperlink
    cell_url = ws1.cell(row=2, column=7)
    assert cell_url.hyperlink is not None or "careers.google.com" in str(cell_url.value)

    # Validate Sheet 2
    ws2 = wb["Skills & ATS Keywords"]
    assert ws2.cell(row=1, column=1).value == "Company"
    assert ws2.cell(row=2, column=1).value == "Google"
    assert "Python" in ws2.cell(row=2, column=3).value
